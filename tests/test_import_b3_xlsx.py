from __future__ import annotations

from pathlib import Path

import pytest


def _make_b3_xlsx(tmp_path: Path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Acoes ---
    ws = wb.create_sheet("Acoes")
    ws.append(
        [
            "Código de Negociação",
            "Quantidade",
            "Preço de Fechamento",
            "Valor Atualizado",
        ]
    )
    ws.append(["AURE3", 10, 12.34, 123.4])
    ws.append(["CXSE3", 5, 10.00, 50.0])
    ws.append(["TAEE3", 2, 40.00, 80.0])

    # --- Fundo de Investimento ---
    ws = wb.create_sheet("Fundo de Investimento")
    ws.append(
        [
            "Código de Negociação",
            "Quantidade",
            "Preço de Fechamento",
            "Valor Atualizado",
        ]
    )

    # ABCP11 duplicated -> should aggregate to qty=265 and price=82.0 (weighted by value)
    # total value = 82 * 265 = 21730
    ws.append(["ABCP11", 100, 80.00, 8000.00])
    ws.append(["ABCP11", 165, 83.00, 13730.00])

    # Single rows (also good to cover non-duplicated)
    ws.append(["BRCR11", 249, 47.73, 11884.77])  # 47.73 * 249
    ws.append(["CNES11", 249, 1.85, 460.65])  # 1.85 * 249

    # --- Tesouro Direto ---
    ws = wb.create_sheet("Tesouro Direto")
    ws.append(["Código ISIN", "Quantidade", "Valor Atualizado"])
    ws.append(["BRSTNCNTB3E2", 24.2, 56575.71])

    out = tmp_path / "posicao.xlsx"
    wb.save(out)
    return out


def test_import_b3_xlsx_parses_positions_and_prices(tmp_path: Path):
    from portfolio_rebalancer.importers.b3_xlsx import import_b3_xlsx

    xlsx = _make_b3_xlsx(tmp_path)
    res = import_b3_xlsx(xlsx, user_id="u1")

    by_ticker = {p.ticker: p for p in res.positions}

    # Stocks
    assert by_ticker["AURE3"].asset_type == "STOCK"
    assert by_ticker["CXSE3"].asset_type == "STOCK"
    assert by_ticker["TAEE3"].asset_type == "STOCK"

    # FIIs aggregated
    assert by_ticker["ABCP11"].asset_type == "FII"
    assert by_ticker["ABCP11"].quantity == pytest.approx(265.0)
    assert by_ticker["ABCP11"].price == pytest.approx(82.0)

    assert by_ticker["BRCR11"].quantity == pytest.approx(249.0)
    assert by_ticker["BRCR11"].price == pytest.approx(47.73)

    assert by_ticker["CNES11"].quantity == pytest.approx(249.0)
    assert by_ticker["CNES11"].price == pytest.approx(1.85)

    # Bonds (ISIN as ticker)
    assert by_ticker["BRSTNCNTB3E2"].asset_type == "BOND"
    assert by_ticker["BRSTNCNTB3E2"].quantity == pytest.approx(24.2)
    assert by_ticker["BRSTNCNTB3E2"].price == pytest.approx(56575.71 / 24.2)


def test_import_b3_xlsx_can_skip_tesouro_sheet(tmp_path: Path):
    from portfolio_rebalancer.importers.b3_xlsx import import_b3_xlsx

    xlsx = _make_b3_xlsx(tmp_path)
    res = import_b3_xlsx(xlsx, user_id="u1", include_tesouro=False)

    assert all(p.asset_type != "BOND" for p in res.positions)
