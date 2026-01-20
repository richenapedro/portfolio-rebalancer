from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from ..models import Position


@dataclass(frozen=True, slots=True)
class B3ImportResult:
    """Normalized data parsed from a B3 'Posicao' XLSX export."""

    user_id: str
    positions: list[Position]
    prices: dict[str, float]
    warnings: list[str]


def import_b3_xlsx(
    path: str | Path,
    *,
    user_id: str,
    include_tesouro: bool = True,
) -> B3ImportResult:
    """Parse a B3 portfolio XLSX and return normalized positions + prices.

    Sheets expected:
    - 'Acoes'
    - 'Fundo de Investimento'
    - 'Tesouro Direto' (optional)

    Notes:
    - Duplicated tickers across institutions/accounts are aggregated (sum of quantities).
    - Price for aggregated positions is computed as weighted average by market value.
    - For Tesouro Direto, we use the ISIN code as the ticker and compute a unit price as
      (Valor Atualizado / Quantidade).
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)

    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(p, data_only=True)

    warnings: list[str] = []

    # key: (asset_type, ticker) -> [qty_sum, value_sum]
    agg: dict[tuple[str, str], list[float]] = {}

    def add_position(asset_type: str, ticker: str, qty: float, value: float) -> None:
        if qty <= 0:
            return
        k = (asset_type, ticker)
        if k not in agg:
            agg[k] = [0.0, 0.0]
        agg[k][0] += qty
        agg[k][1] += value

    # --- Acoes ---
    if "Acoes" in wb.sheetnames:
        _parse_equities_like_sheet(
            wb["Acoes"],
            asset_type="STOCK",
            add_fn=add_position,
            warnings=warnings,
        )
    else:
        warnings.append("Missing sheet 'Acoes'.")

    # --- Fundo de Investimento (FIIs) ---
    if "Fundo de Investimento" in wb.sheetnames:
        _parse_equities_like_sheet(
            wb["Fundo de Investimento"],
            asset_type="FII",
            add_fn=add_position,
            warnings=warnings,
        )
    else:
        warnings.append("Missing sheet 'Fundo de Investimento'.")

    # --- Tesouro Direto ---
    if include_tesouro:
        if "Tesouro Direto" in wb.sheetnames:
            _parse_tesouro_sheet(
                wb["Tesouro Direto"],
                add_fn=add_position,
                warnings=warnings,
            )
        else:
            warnings.append("Missing sheet 'Tesouro Direto'.")

    positions: list[Position] = []
    prices: dict[str, float] = {}

    for (asset_type, ticker), (qty, total_value) in sorted(agg.items()):
        price = (total_value / qty) if qty > 0 else 0.0
        pos = Position(ticker=ticker, asset_type=asset_type, quantity=qty, price=price)
        positions.append(pos)
        prices[pos.ticker] = pos.price

    return B3ImportResult(
        user_id=user_id, positions=positions, prices=prices, warnings=warnings
    )


def _parse_equities_like_sheet(
    ws: Any,
    *,
    asset_type: str,
    add_fn: Any,
    warnings: list[str],
) -> None:
    """Parse 'Acoes' and 'Fundo de Investimento' sheets."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = _header_index(header)

    required_cols = [
        "Código de Negociação",
        "Quantidade",
        "Preço de Fechamento",
        "Valor Atualizado",
    ]
    missing = [c for c in required_cols if c not in idx]
    if missing:
        warnings.append(
            f"Sheet '{ws.title}' missing columns: {missing}. Skipping this sheet."
        )
        return

    t_i = idx["Código de Negociação"]
    q_i = idx["Quantidade"]
    p_i = idx["Preço de Fechamento"]
    v_i = idx["Valor Atualizado"]

    for r in ws.iter_rows(min_row=2, values_only=True):
        ticker = r[t_i]
        if not ticker or not str(ticker).strip():
            continue
        t = str(ticker).strip().upper()

        qty = _coerce_float(r[q_i])
        if qty <= 0:
            continue

        value = _coerce_float(r[v_i])
        if value <= 0:
            price = _coerce_float(r[p_i])
            value = qty * price

        add_fn(asset_type, t, qty, value)


def _parse_tesouro_sheet(
    ws: Any,
    *,
    add_fn: Any,
    warnings: list[str],
) -> None:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = _header_index(header)

    required_cols = ["Código ISIN", "Quantidade", "Valor Atualizado"]
    missing = [c for c in required_cols if c not in idx]
    if missing:
        warnings.append(
            f"Sheet '{ws.title}' missing columns: {missing}. Skipping this sheet."
        )
        return

    isin_i = idx["Código ISIN"]
    qty_i = idx["Quantidade"]
    val_i = idx["Valor Atualizado"]

    for r in ws.iter_rows(min_row=2, values_only=True):
        isin = r[isin_i]
        if not isin or not str(isin).strip():
            continue
        ticker = str(isin).strip().upper()

        qty = _coerce_float(r[qty_i])
        value = _coerce_float(r[val_i])
        if qty <= 0 or value <= 0:
            continue

        add_fn("BOND", ticker, qty, value)


def _header_index(header_row: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip()
        if key:
            out[key] = i
    return out


def _coerce_float(v: Any) -> float:
    """Best-effort conversion to float for common B3 export cell formats."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if not s or s == "-":
        return 0.0

    # pt-BR style: 1.234,56
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except ValueError:
        return 0.0
