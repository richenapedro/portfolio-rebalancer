# .../api/routers/portfolio_db.py
from __future__ import annotations

from io import BytesIO

import openpyxl  # type: ignore
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from ..db.sqlite_db import (
    add_import_run,
    create_portfolio,
    delete_portfolio,
    get_portfolio,
    init_db,
    list_import_runs,
    list_portfolios,
    list_positions,
    rename_portfolio,
    replace_positions,
)
from ..services.auth import CurrentUser, get_current_user
from ..settings import PORTFOLIO_DB_PATH

router = APIRouter(prefix="/api/db", tags=["db"])


class PortfolioCreate(BaseModel):
    name: str = Field(min_length=1, max_length=120)


class PortfolioRename(BaseModel):
    name: str = Field(min_length=1, max_length=120)


class PositionIn(BaseModel):
    ticker: str
    quantity: float
    price: float | None = None
    cls: str | None = None
    note: int | None = None
    source: str = "manual"  # import/manual


class ReplacePositionsBody(BaseModel):
    positions: list[PositionIn]


class ImportRunCreate(BaseModel):
    filename: str = Field(min_length=1, max_length=255)


@router.get("/health")
def health(user: CurrentUser = Depends(get_current_user)):
    init_db(PORTFOLIO_DB_PATH)
    return {"ok": True, "db_path": PORTFOLIO_DB_PATH, "user_id": user.id}


@router.get("/portfolios")
def portfolios_list(user: CurrentUser = Depends(get_current_user)):
    return {"items": list_portfolios(PORTFOLIO_DB_PATH, user.id)}


@router.post("/portfolios")
def portfolios_create(body: PortfolioCreate, user: CurrentUser = Depends(get_current_user)):
    return create_portfolio(PORTFOLIO_DB_PATH, user.id, body.name)


@router.get("/portfolios/{portfolio_id}")
def portfolios_get(portfolio_id: int, user: CurrentUser = Depends(get_current_user)):
    p = get_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")
    return p


@router.put("/portfolios/{portfolio_id}")
def portfolios_rename(
    portfolio_id: int, body: PortfolioRename, user: CurrentUser = Depends(get_current_user)
):
    p = rename_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id, body.name)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")
    return p


@router.get("/portfolios/{portfolio_id}/positions")
def positions_list(portfolio_id: int, user: CurrentUser = Depends(get_current_user)):
    p = get_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")
    return {"items": list_positions(PORTFOLIO_DB_PATH, user.id, portfolio_id)}


@router.post("/portfolios/{portfolio_id}/positions/replace")
def positions_replace(
    portfolio_id: int, body: ReplacePositionsBody, user: CurrentUser = Depends(get_current_user)
):
    p = get_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")

    payload = [x.model_dump() for x in body.positions]
    replace_positions(PORTFOLIO_DB_PATH, user.id, portfolio_id, payload)
    return {"ok": True}


@router.get("/portfolios/{portfolio_id}/import_runs")
def import_runs_list(portfolio_id: int, user: CurrentUser = Depends(get_current_user)):
    p = get_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")
    return {"items": list_import_runs(PORTFOLIO_DB_PATH, user.id, portfolio_id)}


@router.post("/portfolios/{portfolio_id}/import_runs")
def import_runs_add(
    portfolio_id: int, body: ImportRunCreate, user: CurrentUser = Depends(get_current_user)
):
    p = get_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")
    return add_import_run(PORTFOLIO_DB_PATH, user.id, portfolio_id, body.filename)


def _classify_db_row(ticker: str, cls: str | None) -> str:
    c = (cls or "").strip().lower()
    t = (ticker or "").strip().upper()

    if c in {"fiis", "fii"}:
        return "FII"
    if c in {"bonds", "bond", "tesouro", "rf"}:
        return "BOND"
    if c in {"stocks", "stock", "acoes", "ação", "acoes"}:
        return "STOCK"

    if t.endswith("11"):
        return "FII"
    if t.startswith("BRSTN"):
        return "BOND"
    return "STOCK"


@router.get("/portfolios/{portfolio_id}/export_b3_xlsx")
def export_b3_xlsx(portfolio_id: int, user: CurrentUser = Depends(get_current_user)):
    p = get_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")

    rows = list_positions(PORTFOLIO_DB_PATH, user.id, portfolio_id)

    wb = openpyxl.Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_acoes = wb.create_sheet("Acoes")
    ws_fiis = wb.create_sheet("Fundo de Investimento")
    ws_td = wb.create_sheet("Tesouro Direto")

    ws_acoes.append(["Código de Negociação", "Quantidade", "Preço de Fechamento", "Valor Atualizado"])
    ws_fiis.append(["Código de Negociação", "Quantidade", "Preço de Fechamento", "Valor Atualizado"])
    ws_td.append(["Código ISIN", "Quantidade", "Valor Atualizado"])

    for r in rows:
        ticker = str(r.get("ticker") or "").strip().upper()
        if not ticker:
            continue

        qty = float(r.get("quantity") or 0.0)
        if qty <= 0:
            continue

        price = r.get("price")
        price_f = float(price) if price is not None else 0.0
        value = qty * price_f

        kind = _classify_db_row(ticker, r.get("cls"))

        if kind == "FII":
            ws_fiis.append([ticker, qty, price_f, value])
        elif kind == "BOND":
            ws_td.append([ticker, qty, value])
        else:
            ws_acoes.append([ticker, qty, price_f, value])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"portfolio_{portfolio_id}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/portfolios/{portfolio_id}")
def portfolios_delete(portfolio_id: int, user: CurrentUser = Depends(get_current_user)):
    p = get_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    if not p:
        raise HTTPException(status_code=404, detail="Portfolio não encontrado.")
    delete_portfolio(PORTFOLIO_DB_PATH, user.id, portfolio_id)
    return {"ok": True}